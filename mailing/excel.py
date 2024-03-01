import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

from accounts.models import Owner

from bot.apps import BotConfig
from telebot.apihelper import ApiTelegramException

bot = BotConfig.bot


class ExcelDebtsMailing:

    def __init__(self, name):
        self.wb = openpyxl.load_workbook(name)
        self.ws = self.wb.worksheets[0]

    def _check_debt(self, delay, amount):
        if amount < 10:
            return False
        if -6 < delay <= 0:
            return amount >= 50
        if -300 < delay <= -6:
            return amount > 10

    def _generate_message(self, amount):
        return (f"Добрый день! Напоминаем, что <b>Ваша ПРОСРОЧЕННАЯ ЗАДОЛЖЕННОСТЬ "
                f"составляет {amount:.2f} руб.</b> Просим погасить сегодня-завтра")

    def mailing(self):
        self._set_debts()
        for unp, amount in self.debts.items():
            owner = Owner.objects.filter(unp=unp).first()

            if not owner:
                continue
            if not owner.tg_id or not owner.has_notifications:
                continue

            try:
                bot.send_message(owner.tg_id, self._generate_message(amount), parse_mode="HTML")
            except ApiTelegramException:
                print(f"Вероятней всего у пользователя неправильный id телеграма. Проверьте {owner}")

    def _set_debts(self):
        row_index = 7
        self.debts = dict()
        while True:
            unp = self.ws[f"G{row_index}"].value
            if not unp:
                break

            delay = self.ws[f"I{row_index}"].value
            amount = self.ws[f"J{row_index}"].value

            if not (delay and amount) or not (isinstance(delay, int) and isinstance(amount, float)):
                row_index += 1
                continue

            if self._check_debt(delay=delay, amount=amount):
                self.debts[unp] = amount

            row_index += 1

