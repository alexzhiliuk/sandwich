import openpyxl

from django.db.utils import IntegrityError
from accounts.models import Owner

from bot.apps import BotConfig

bot = BotConfig.bot


class ExcelRegistrateUsers:

    def __init__(self, file):
        self.wb = openpyxl.load_workbook(file)
        self.ws = self.wb.worksheets[0]

    def register(self):
        row_index = 3

        while True:

            unp = self.ws[f"B{row_index}"].value
            if unp is None:
                break

            owner = Owner.objects.filter(unp=unp)
            if owner:
                row_index += 1
                continue

            name = self.ws[f"A{row_index}"].value
            phone = self.ws[f"C{row_index}"].value

            if name and phone:
                try:
                    Owner.objects.create(unp=unp, fio=name, phone=phone, is_active=True)
                except IntegrityError:
                    pass

            row_index += 1







