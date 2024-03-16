import datetime

import openpyxl
from openpyxl import Workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.fills import PatternFill
from openpyxl.styles import NamedStyle, Alignment, Font

from orders.models import Product, ProductType, Order


class ExcelDailyReport:
    cells_indexes = {
        "ФИО": 1,
        "УНП": 2,
        "000000039": 3,
        "000000036": 5,
        "000000280": 7,
        "000000038": 9,
        "000000397": 11,
        "000000399": 13,
        "000000245": 15,
        "000000246": 17,
        "000000259": 19,
        "000000400": 21,
        "000000402": 23,
        "000000401": 25,
        "000000404": 27,
        "000000406": 29,
        "000000405": 31,
        "000000403": 33,
        "Итого": 35,
        "Пункт разгрузки и примечание": 36,
    }

    def __init__(self, name, date):
        self.wb = openpyxl.load_workbook(name)
        self.ws = self.wb.worksheets[0]
        self.date = date

        self.formula_fields = [3, 5, 7, 9, 11, 13, 15, 17, 19, 21, 23, 25, 27, 29, 31, 33, 35]

        self._create_styles()
        self._write()

        for col in self.ws.iter_cols():
            for cell in col:
                cell.border = self.thin_border

    def _create_styles(self):
        self.medium_border = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'),
                               bottom=Side(style='medium'))
        self.thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                    bottom=Side(style='thin'))

        self.fillers = [
            PatternFill(patternType="solid", fgColor="00FFCC99"),
            PatternFill(patternType="solid", fgColor="0099CCFF"),
            PatternFill(patternType="solid", fgColor="00CCFFCC")
        ]

        self.product_type_bg = {
            prod_type: self.fillers[i] for i, prod_type in enumerate(ProductType.objects.all()[:3])
        }

        self.title_font = Font(name="Arial", size=10, bold=True)
        self.count_font = Font(name="Arial", size=9, bold=True)
        self.price_font = Font(name="Arial", size=9, italic=True)
        self.default_font = Font(name="Arial", size=9)

        self.center_alignment = Alignment(horizontal='center')

    def _fill_in_order_row_simple_info(self, order):
        self.row["ФИО"] = order.owner.fio
        self.row["УНП"] = order.owner.unp

        if order.point:
            self.row["Пункт разгрузки и примечание"] = order.point.address
        else:
            self.row["Пункт разгрузки и примечание"] = "Самовывоз"

    def _fill_in_order_row_products_info(self, order):
        ordered_products = set()
        for item in order.items.all():
            count = item.count
            price = item.price

            self.row[item.product] = (count, round(price / count, 2))
            ordered_products.add(item.product)

        for product in set(Product.objects.all()).difference(ordered_products):
            special_price = product.get_special_price_for_user(order.owner)
            if special_price:
                self.row[product] = (None, special_price.price)
            else:
                self.row[product] = (None, product.price)

    def _fill_in_order_row_summarize(self, row_index):
        sum_fields = []
        for col_num in self.formula_fields[:-1]:
            col_letter = self.ws.cell(1, col_num).column_letter
            sum_fields.append(f"{col_letter}{row_index}")
        self.row["Итого"] = f"={'+'.join(sum_fields)}"

    def _get_order_row(self, order, row_index):
        self.row = dict()

        self._fill_in_order_row_simple_info(order)
        self._fill_in_order_row_products_info(order)
        self._fill_in_order_row_summarize(row_index)

        return self.row

    def _fill_in_orders(self):

        self.orders = Order.objects.filter(created_at__date=self.date)

        row_index = 2
        for order in self.orders:
            row = self._get_order_row(order, row_index)
            for col, value in row.items():

                if isinstance(col, Product):
                    code = col.excel_code
                    count_cell = self.ws.cell(row_index, self.cells_indexes[code], value[0])
                    price_cell = self.ws.cell(row_index, self.cells_indexes[code] + 1, value[1])
                    count_cell.font = self.count_font
                    count_cell.alignment = self.center_alignment
                    price_cell.font = self.price_font
                    price_cell.alignment = self.center_alignment
                else:
                    cell = self.ws.cell(row_index, self.cells_indexes[col], value)
                    if col == "Итого":
                        cell.font = self.title_font
                        cell.alignment = self.center_alignment
                    else:
                        cell.font = self.default_font

            row_index += 1

        return row_index

    def _summarize(self, row_index):
        # Итого внизу таблицы
        self.ws.cell(row_index + 1, 1, "ИТОГО:")
        for col_num in self.formula_fields:
            col_letter = self.ws.cell(1, col_num).column_letter

            cell = self.ws.cell(row_index + 1, col_num, f"=SUM({col_letter}1:{col_letter}{row_index})")
            cell.font = self.title_font
            cell.alignment = self.center_alignment

    def _write(self):
        row_index = self._fill_in_orders()
        self._summarize(row_index)


class ExcelDriverReport(ExcelDailyReport):
    cells_indexes = {
        "ФИО": 1,
        "УНП": 2,
        "000000039": 3,
        "000000036": 5,
        "000000280": 7,
        "000000038": 9,
        "000000397": 11,
        "000000399": 13,
        "000000245": 15,
        "000000246": 17,
        "000000259": 19,
        "000000400": 21,
        "000000402": 23,
        "000000401": 25,
        "000000404": 27,
        "000000406": 29,
        "000000405": 31,
        "000000403": 33,
        "Итого": 35,
        "Пункт разгрузки и примечание": 36,
        "Телефон": 37,
        "Время": 38,
        "Сумма": 39,
    }

    def _fill_in_order_row_simple_info(self, order):
        self.row["ФИО"] = order.owner.fio
        self.row["УНП"] = order.owner.unp

        if order.point:
            self.row["Пункт разгрузки и примечание"] = order.point.address
        else:
            self.row["Пункт разгрузки и примечание"] = "Самовывоз"

        if order.employee:
            self.row["Телефон"] = order.employee.phone
        else:
            self.row["Телефон"] = order.owner.phone

        self.row["Время"] = order.point.working_hours or "" if order.point else ""
        self.row["Сумма"] = order.get_final_info()[1]

    def _write(self):
        self._fill_in_orders()


class ExcelLabelmakerReport:
    cells_indexes = {
        "ФИО": 1,
        "000000039": 2,
        "000000036": 3,
        "000000280": 4,
        "000000038": 5,
        "000000397": 6,
        "000000399": 7,
        "000000245": 8,
        "000000246": 9,
        "000000259": 10,
        "000000400": 11,
        "000000402": 12,
        "000000401": 13,
        "000000404": 14,
        "000000406": 15,
        "000000405": 16,
        "000000403": 17,
        "Итого": 18,
        "Адрес": 19,
        "Телефон": 20,
        "Дата": 21,
    }
    default_cells = {
        22: "Кур", 23: "Вет", 24: "Бек", 25: "Рыб", 26: "Сыт",
        27: "Гав", 28: "Ркур", 29: "Рвет", 30: "Рсос", 31: "Рпик",
        32: "Б.кур", 33: "Б.вет", 34: "Б.рыб", 35: "Б.сгу", 36: "Б.тво",
        37: "Б.дже", 38: "Итого"
    }
    calendar = {1: "Января", 2: "Февраля",  3: "Марта",  4: "Апреля",  5: "Мая",  6: "Июня",
                7: "Июля",  8: "Августа",  9: "Сентября",  10: "Октября",  11: "Ноября", 12: "Декабря"}
    products = Product.objects.all()

    def __init__(self, name, date):
        self.wb = openpyxl.load_workbook(name)
        self.ws = self.wb.worksheets[0]
        self.date = date

        self._create_styles()
        self._write()

    def _create_styles(self):
        self.medium_border = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'),
                                    bottom=Side(style='medium'))
        self.thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                  bottom=Side(style='thin'))

        self.red_bg = PatternFill(patternType="solid", fgColor="00FF0000")

        self.title_font = Font(name="Arial", size=10, bold=True)
        self.count_font = Font(name="Arial", size=9, bold=True)
        self.price_font = Font(name="Arial", size=9, italic=True)
        self.default_font = Font(name="Arial", size=9)

        self.center_alignment = Alignment(horizontal='center')

        self.red_cell = NamedStyle(name="red_cell", fill=self.red_bg, border=self.medium_border)
        self.simple_cell = NamedStyle(name="simple_cell", border=self.medium_border)
        self.wb.add_named_style(self.red_cell)
        self.wb.add_named_style(self.simple_cell)

    def _get_row_simple_info(self, order):
        self.row["ФИО"] = order.owner.fio
        self.row["Адрес"] = order.point.address if order.point else "Самовывоз"
        self.row["Телефон"] = order.employee.phone if order.employee else order.owner.phone

        tomorrow = datetime.date.today() + datetime.timedelta(1)
        self.row["Дата"] = f"{tomorrow.day} {self.calendar[tomorrow.month]}"

    def _get_row_products_info(self, order, row_index):

        ordered_products = set()
        items = order.items.all()
        for item in items:
            self.row[f"{item.product.excel_code}"] = item.count
            ordered_products.add(item.product)

        for product in set(self.products).difference(ordered_products):
            self.row[f"{product.excel_code}"] = ""

        start_col_letter = self.ws.cell(1, 2).column_letter
        end_col_letter = self.ws.cell(1, self.cells_indexes["Итого"] - 1).column_letter
        self.row["Итого"] = f"=SUM({start_col_letter}{row_index}:{end_col_letter}{row_index})"

    def _get_row_info(self, order, row_index):
        self.row = dict()
        self._get_row_simple_info(order)
        self._get_row_products_info(order, row_index)

        return self.row

    def _add_default_info(self, row_index):

        for col, value in self.default_cells.items():
            cell = self.ws.cell(row_index, col, value)
            cell.style = self.red_cell

    def _fill_in(self):
        orders = Order.objects.filter(created_at__date=self.date)

        row_index = 2
        for order in orders:
            row_info = self._get_row_info(order, row_index)

            for col, value in row_info.items():
                col_index = self.cells_indexes[col]
                cell = self.ws.cell(row_index, col_index, value)

                if col == "Дата":
                    cell.style = self.red_cell
                else:
                    cell.style = self.simple_cell

            self._add_default_info(row_index)

            row_index += 1

    def _write(self):
        self._fill_in()